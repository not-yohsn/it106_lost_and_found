"""Unit tests for the Excel + PDF export helpers."""
import io

from openpyxl import load_workbook

from app.exports import build_xlsx, render_pdf


def test_build_xlsx_returns_loadable_workbook_with_header_and_rows():
    headers = ["ID", "Name", "Status"]
    rows = [
        (1, "Black Backpack", "matched"),
        (2, "Silver Pen", "reported"),
    ]
    payload = build_xlsx("Test Sheet", headers, rows)

    wb = load_workbook(io.BytesIO(payload))
    ws = wb.active
    assert ws.title == "Test Sheet"
    assert [cell.value for cell in ws[1]] == headers
    assert [cell.value for cell in ws[2]] == [1, "Black Backpack", "matched"]
    assert [cell.value for cell in ws[3]] == [2, "Silver Pen", "reported"]


def test_render_pdf_returns_pdf_magic_bytes():
    html = "<html><body><h1>Receipt</h1><p>Hello</p></body></html>"
    payload = render_pdf(html)
    assert payload[:4] == b"%PDF"
    assert len(payload) > 500  # any real PDF will be larger than this
