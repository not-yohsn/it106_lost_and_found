"""Export helpers — Excel via openpyxl and PDF via xhtml2pdf.

Keeping the spreadsheet/PDF building logic out of the route modules so the
routes stay focused on auth, query, and response.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from xhtml2pdf import pisa


def _autosize(ws):
    """Cheap column-width autosize so the exported xlsx is readable."""
    for column_cells in ws.columns:
        length = max((len(str(c.value)) for c in column_cells if c.value is not None), default=10)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 60)


def _write_header(ws, headers):
    bold_white = Font(bold=True, color="FFFFFF")
    indigo_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    centered = Alignment(horizontal="center")
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = bold_white
        cell.fill = indigo_fill
        cell.alignment = centered


def build_xlsx(sheet_name, headers, rows):
    """Build a single-sheet workbook in memory and return the bytes.

    `rows` is an iterable of sequences whose length matches `headers`.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel limits sheet names to 31 chars
    _write_header(ws, headers)
    for row in rows:
        ws.append(list(row))
    _autosize(ws)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def render_pdf(html):
    """Render an HTML string into a PDF (bytes) using xhtml2pdf.

    Pure-Python — no GTK / wkhtmltopdf binary required, which makes it
    portable to Windows / school laptops.
    """
    buffer = io.BytesIO()
    result = pisa.CreatePDF(src=html, dest=buffer)
    if result.err:
        raise RuntimeError("PDF rendering failed")
    buffer.seek(0)
    return buffer.read()
